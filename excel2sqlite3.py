#import tkinter
import tkinter as tk
#import ttk module which has the notebook widget                     
from tkinter import ttk
#import all from tkinter module, StringVar is in it
from tkinter import *
#import openpyxl module for interacting with excel spreadsheets and sqlite3 for the database
import openpyxl, sqlite3


#Function which contains everything related to automatic entry of data
def unos_podataka(event):
    #takes the path to the excel file from the Entry widget
    putanja_text = putanja_automatski_unos.get()
    #openpyxl access data from the excel spreadsheet
    workbook = openpyxl.load_workbook(putanja_text)
    #work with the active sheet
    sheet = workbook.active
    
    #read data directly from the excel spreadsheet
    #left side of excel spreadsheet
    #define Label, tab which contains it, label text, add the value from excel, grid positions it on the frame, sticky for a cleaner look
    djelatnik = Label(tab1, text=f"Djelatnik: {sheet.cell(row=5, column=3).value}").grid(row=1, column=0, sticky="E")
    naziv_racunala = Label(tab1, text=f"Naziv računala: {sheet.cell(row=6, column=3).value}").grid(row=2, column=0, sticky="E")
    korisnicko_ime = Label(tab1, text=f"Korisničko ime: {sheet.cell(row=7, column=3).value}").grid(row=3, column=0, sticky="E")
    lozinka = Label(tab1, text=f"Lozinka: {sheet.cell(row=8, column=3).value}").grid(row=4, column=0, sticky="E")
    inventurni_broj = Label(tab1, text=f"Inventurni broj: {sheet.cell(row=9, column=3).value}").grid(row=5, column=0, sticky="E")
    mac_adresa = Label(tab1, text=f"MAC adresa: {sheet.cell(row=10, column=3).value}").grid(row=6, column=0, sticky="E")
    operacijski_sustav = Label(tab1, text=f"Operacijski sustav: {sheet.cell(row=14, column=3).value}").grid(row=7, column=0, sticky="E")
    dodatni_program1 = Label(tab1, text=f"Dodatni program 1: {sheet.cell(row=16, column=2).value}").grid(row=8, column=0, sticky="E")
    dodatni_program2 = Label(tab1, text=f"Dodatni program 2: {sheet.cell(row=17, column=2).value}").grid(row=9, column=0, sticky="E")
    dodatni_program3 = Label(tab1, text=f"Dodatni program 3: {sheet.cell(row=18, column=2).value}").grid(row=10, column=0, sticky="E")
    dodatni_program4 = Label(tab1, text=f"Dodatni program 4: {sheet.cell(row=19, column=2).value}").grid(row=11, column=0, sticky="E")
    dodatni_program5 = Label(tab1, text=f"Dodatni program 5: {sheet.cell(row=20, column=2).value}").grid(row=12, column=0, sticky="E")
    dodatni_program6 = Label(tab1, text=f"Dodatni program 6: {sheet.cell(row=21, column=2).value}").grid(row=13, column=0, sticky="E")
    dodatni_program7 = Label(tab1, text=f"Dodatni program 7: {sheet.cell(row=22, column=2).value}").grid(row=14, column=0, sticky="E")
    dodatni_program8 = Label(tab1, text=f"Dodatni program 8: {sheet.cell(row=23, column=2).value}").grid(row=15, column=0, sticky="E")
    dodatni_program9 = Label(tab1, text=f"Dodatni program 9: {sheet.cell(row=24, column=2).value}").grid(row=16, column=0, sticky="E")
    dodatni_program10 = Label(tab1, text=f"Dodatni program 10: {sheet.cell(row=25, column=2).value}").grid(row=17, column=0, sticky="E")
    dodatni_program11 = Label(tab1, text=f"Dodatni program 11: {sheet.cell(row=26, column=2).value}").grid(row=18, column=0, sticky="E")
    dodatni_program12 = Label(tab1, text=f"Dodatni program 12: {sheet.cell(row=27, column=2).value}").grid(row=19, column=0, sticky="E")
    dodatni_program13 = Label(tab1, text=f"Dodatni program 13: {sheet.cell(row=28, column=2).value}").grid(row=20, column=0, sticky="E")
    dodatni_program14 = Label(tab1, text=f"Dodatni program 14: {sheet.cell(row=29, column=2).value}").grid(row=21, column=0, sticky="E")
    dodatni_program15 = Label(tab1, text=f"Dodatni program 15: {sheet.cell(row=30, column=2).value}").grid(row=22, column=0, sticky="E")
    
    #right side of excel spreadsheet
    model = Label(tab1, text=f"Model: {sheet.cell(row=5, column=6).value}").grid(row=1, column=1, sticky="E")
    cpu = Label(tab1, text=f"CPU: {sheet.cell(row=6, column=6).value}").grid(row=2, column=1, sticky="E")
    ram = Label(tab1, text=f"RAM: {sheet.cell(row=7, column=6).value}").grid(row=3, column=1, sticky="E")
    mrezna_kartica = Label(tab1, text=f"Mrežna kartica: {sheet.cell(row=8, column=6).value}").grid(row=4, column=1, sticky="E")
    graficka_kartica = Label(tab1, text=f"Grafička kartica: {sheet.cell(row=9, column=6).value}").grid(row=5, column=1, sticky="E")
    glavna_particija = Label(tab1, text=f"Glavna particija: {sheet.cell(row=10, column=6).value}").grid(row=6, column=1, sticky="E")
    dodatne_particije = Label(tab1, text=f"Dodatne particije: {sheet.cell(row=11, column=6).value}").grid(row=7, column=1, sticky="E")
    napomena1 = Label(tab1, text=f"Napomena 1: {sheet.cell(row=14, column=5).value}").grid(row=8, column=1, sticky="E")
    napomena2 = Label(tab1, text=f"Napomena 2: {sheet.cell(row=15, column=5).value}").grid(row=9, column=1, sticky="E")
    napomena3 = Label(tab1, text=f"Napomena 3: {sheet.cell(row=16, column=5).value}").grid(row=10, column=1, sticky="E")
    napomena4 = Label(tab1, text=f"Napomena 4: {sheet.cell(row=17, column=5).value}").grid(row=11, column=1, sticky="E")
    napomena5 = Label(tab1, text=f"Napomena 5: {sheet.cell(row=18, column=5).value}").grid(row=12, column=1, sticky="E")
    napomena6 = Label(tab1, text=f"Napomena 6: {sheet.cell(row=19, column=5).value}").grid(row=13, column=1, sticky="E")
    napomena7 = Label(tab1, text=f"Napomena 7: {sheet.cell(row=20, column=5).value}").grid(row=14, column=1, sticky="E")
    napomena8 = Label(tab1, text=f"Napomena 8: {sheet.cell(row=21, column=5).value}").grid(row=15, column=1, sticky="E")
    napomena9 = Label(tab1, text=f"Napomena 9: {sheet.cell(row=22, column=5).value}").grid(row=16, column=1, sticky="E")
    napomena10 = Label(tab1, text=f"Napomena 10: {sheet.cell(row=23, column=5).value}").grid(row=17, column=1, sticky="E")
    napomena11 = Label(tab1, text=f"Napomena 11: {sheet.cell(row=24, column=5).value}").grid(row=18, column=1, sticky="E")
    napomena12 = Label(tab1, text=f"Napomena 12: {sheet.cell(row=25, column=5).value}").grid(row=19, column=1, sticky="E")
    napomena13 = Label(tab1, text=f"Napomena 13: {sheet.cell(row=26, column=5).value}").grid(row=20, column=1, sticky="E")
    napomena14 = Label(tab1, text=f"Napomena 14: {sheet.cell(row=27, column=5).value}").grid(row=21, column=1, sticky="E")
    napomena15 = Label(tab1, text=f"Napomena 15: {sheet.cell(row=28, column=5).value}").grid(row=22, column=1, sticky="E")
    napomena16 = Label(tab1, text=f"Napomena 16: {sheet.cell(row=29, column=5).value}").grid(row=23, column=1, sticky="E")
    napomena17 = Label(tab1, text=f"Napomena 17: {sheet.cell(row=30, column=5).value}").grid(row=24, column=1, sticky="E")

    #date and employee who did the work
    radove_izveo = Label(tab1, text=f"Radove izveo: {sheet.cell(row=32, column=6).value}").grid(row=31, column=0, sticky="E")
    datum = Label(tab1, text=f"Datum: {sheet.cell(row=32, column=3).value}").grid(row=32, column=0, sticky="E")
    
    #connect to database with the path we specify, or create a new database at that location if it does not exist
    #define proper path for your own machine
    conn = sqlite3.connect('/home/user/database.db')
    #cursor is needed for browsing data in the database, without it we can only enter data into the database
    c = conn.cursor()

    #create the table if it doesn't exist, with the entries we need
    # we have to define the data type which we will enter in specified fields
    # SQL commands in capital letters, rest as you will 
    def stvori_tablicu_automatski():
        c.execute("CREATE TABLE IF NOT EXISTS izvjestaj_tablica(inventurni_broj INTEGER, djelatnik TEXT, "
                  "naziv_racunala TEXT, korisnicko_ime TEXT, lozinka TEXT, mac_adresa TEXT, operacijski_sustav TEXT, "
                  "dodatni_program1 TEXT, dodatni_program2 TEXT, dodatni_program3 TEXT, dodatni_program4 TEXT, "
                  "dodatni_program5 TEXT, dodatni_program6 TEXT, dodatni_program7 TEXT, dodatni_program8 TEXT, "
                  "dodatni_program9 TEXT, dodatni_program10 TEXT, dodatni_program11 TEXT, dodatni_program12 TEXT, "
                  "dodatni_program13 TEXT, dodatni_program14 TEXT, dodatni_program15 TEXT, model TEXT, cpu TEXT, "
                  "ram TEXT, mrezna_kartica TEXT, graficka_kartica TEXT, glavna_particija TEXT, dodatne_particije TEXT, "
                  "napomena1 TEXT, napomena2 TEXT, napomena3 TEXT, napomena4 TEXT, napomena5 TEXT, napomena6 TEXT, "
                  "napomena7 TEXT, napomena8 TEXT, napomena9 TEXT, napomena10 TEXT, napomena11 TEXT, napomena12 TEXT, "
                  "napomena13 TEXT, napomena14 TEXT, napomena15 TEXT, napomena16 TEXT, napomena17 TEXT, datum TEXT, "
                  "radove_izveo TEXT)")
    
    #after creating the table we define exactly what we are entering into it
    #definition format: (tablica(stupac),?)
    #in this case e.g. (izvještaj_tablica(djelatnik),?),sheet.cell(row=5, column=3).value to pull directly from excel, openpyxl is used here
    def unesi_podatke_automatski():
        c.execute("INSERT INTO izvjestaj_tablica(inventurni_broj, djelatnik, naziv_racunala, korisnicko_ime, lozinka, "
                  "mac_adresa, operacijski_sustav, dodatni_program1, dodatni_program2, dodatni_program3, dodatni_program4, "
                  "dodatni_program5, dodatni_program6, dodatni_program7, dodatni_program8, dodatni_program9, "
                  "dodatni_program10, dodatni_program11, dodatni_program12, dodatni_program13, dodatni_program14, "
                  "dodatni_program15, model, cpu, ram, mrezna_kartica, graficka_kartica, glavna_particija, "
                  "dodatne_particije, napomena1, napomena2, napomena3, napomena4, napomena5, napomena6, napomena7, "
                  "napomena8, napomena9, napomena10, napomena11, napomena12, napomena13, napomena14, napomena15, "
                  "napomena16, napomena17, datum, radove_izveo) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, "
                  "?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, "
                  "?)", (sheet.cell(row=9, column=3).value, sheet.cell(row=5, column=3).value, sheet.cell(row=6, column=3).value, 
                        sheet.cell(row=7, column=3).value, sheet.cell(row=8, column=3).value, sheet.cell(row=10, column=3).value,
                        sheet.cell(row=14, column=3).value, sheet.cell(row=16, column=2).value, sheet.cell(row=17, column=2).value,
                        sheet.cell(row=18, column=2).value, sheet.cell(row=19, column=2).value, sheet.cell(row=20, column=2).value,
                        sheet.cell(row=21, column=2).value, sheet.cell(row=22, column=2).value, sheet.cell(row=23, column=2).value,
                        sheet.cell(row=24, column=2).value, sheet.cell(row=25, column=2).value, sheet.cell(row=26, column=2).value,
                        sheet.cell(row=27, column=2).value, sheet.cell(row=28, column=2).value, sheet.cell(row=29, column=2).value, 
                        sheet.cell(row=30, column=2).value, sheet.cell(row=5, column=6).value, sheet.cell(row=6, column=6).value, 
                        sheet.cell(row=7, column=6).value, sheet.cell(row=8, column=6).value, sheet.cell(row=9, column=6).value, 
                        sheet.cell(row=10, column=6).value, sheet.cell(row=11, column=6).value, sheet.cell(row=14, column=5).value, 
                        sheet.cell(row=15, column=5).value, sheet.cell(row=16, column=5).value, sheet.cell(row=17, column=5).value, 
                        sheet.cell(row=18, column=5).value, sheet.cell(row=19, column=5).value, sheet.cell(row=20, column=5).value, 
                        sheet.cell(row=21, column=5).value, sheet.cell(row=22, column=5).value, sheet.cell(row=23, column=5).value, 
                        sheet.cell(row=24, column=5).value, sheet.cell(row=25, column=5).value, sheet.cell(row=26, column=5).value, 
                        sheet.cell(row=27, column=5).value, sheet.cell(row=28, column=5).value, sheet.cell(row=29, column=5).value, 
                        sheet.cell(row=30, column=5).value, sheet.cell(row=32, column=3).value, sheet.cell(row=32, column=6).value))
        #confirm changes and close connection to the database               
        conn.commit()

    #function that calls all functions related to automatic storing via button "Pohrani u BP!"
    def automatski_upis(event):
        stvori_tablicu_automatski()
        unesi_podatke_automatski()
        #remove button after click
        pohrani_podatke_automatski.destroy()
        #new label instead of the button, so the user knows it's done
        gotovo_automatski_label = Label(tab1, text="Pohranjeno!").grid(row=33, column=0)
        
    #button for storing to database, for executing all defined functions
    pohrani_podatke_automatski = Button(tab1, width=15, height=1)
    pohrani_podatke_automatski.config(text='Pohrani u BP!', fg='red', bg='silver')
    pohrani_podatke_automatski.bind("<Button-1>", automatski_upis)
    pohrani_podatke_automatski.grid(row=33, column=0)


#manual entry
def rucni_unos(event):
    #define variables for manual entry
    djelatnik_rucno = StringVar()
    naziv_racunala_rucno = StringVar()
    korisnicko_ime_rucno = StringVar()
    lozinka_rucno = StringVar()
    inventurni_broj_rucno = StringVar()
    mac_adresa_rucno = StringVar()
    model_rucno = StringVar()
    cpu_rucno = StringVar()
    ram_rucno = StringVar()
    mrezna_kartica_rucno = StringVar()
    graficka_kartica_rucno = StringVar()
    glavna_particija_rucno = StringVar()
    dodatne_particije_rucno = StringVar()
    os_rucno = StringVar()
    napomena1_rucno = StringVar()
    napomena2_rucno = StringVar()
    napomena3_rucno = StringVar()
    napomena4_rucno = StringVar()
    napomena5_rucno = StringVar()
    napomena6_rucno = StringVar()
    napomena7_rucno = StringVar()
    napomena8_rucno = StringVar()
    napomena9_rucno = StringVar()
    napomena10_rucno = StringVar()
    napomena11_rucno = StringVar()
    napomena12_rucno = StringVar()  
    napomena13_rucno = StringVar()
    napomena14_rucno = StringVar()  
    napomena15_rucno = StringVar()
    napomena16_rucno = StringVar()
    napomena17_rucno = StringVar()
    program1_rucno = StringVar()
    program2_rucno = StringVar()
    program3_rucno = StringVar()
    program4_rucno = StringVar()
    program5_rucno = StringVar()
    program6_rucno = StringVar()
    program7_rucno = StringVar()
    program8_rucno = StringVar()
    program9_rucno = StringVar()
    program10_rucno = StringVar()
    program11_rucno = StringVar()
    program12_rucno = StringVar()
    program13_rucno = StringVar()
    program14_rucno = StringVar()
    program15_rucno = StringVar()
    datum_rucno = StringVar()
    radove_izveo_rucno = StringVar()
    operacijski_sustav_rucno = StringVar()
    
    #remove button after click
    rucni_unos_button.destroy()

    #tab2, label and entry for basic info
    #textvariable takes the variables we defined above and stores within them the text we enter into the entry widgets
    #during entry into the database the get() method gets the variable values
    label_informacije = Label(tab2, text="Informacije: ").grid(row=0, column=0, sticky="e")
    label_djelatnik = Label(tab2, text="Djelatnik: ").grid(row=1, column=0, sticky="e")
    entry_djelatnik = Entry(tab2, textvariable=djelatnik_rucno, width = 25).grid(row=1, column=1)
    label_naziv_racunala = Label(tab2, text="Naziv računala: ").grid(row=2, column=0, sticky="e")
    entry_naziv_racunala = Entry(tab2, textvariable=naziv_racunala_rucno, width = 25).grid(row=2, column=1)
    label_korisnicko_ime = Label(tab2, text="Korisničko ime ").grid(row=3, column=0, sticky="e")
    entry_korisnicko_ime = Entry(tab2, textvariable=korisnicko_ime_rucno, width = 25).grid(row=3, column=1)
    label_lozinka = Label(tab2, text="Lozinka ").grid(row=4, column=0, sticky="e")
    entry_lozinka = Entry(tab2, textvariable=lozinka_rucno, width = 25).grid(row=4, column=1)
    label_inventurni_broj = Label(tab2, text="Inventurni broj: ").grid(row=5, column=0, sticky="e")
    entry_inventurni_broj = Entry(tab2, textvariable=inventurni_broj_rucno, width = 25).grid(row=5, column=1)
    label_mac_adresa = Label(tab2, text="MAC adresa: ").grid(row=6, column=0, sticky="e")
    entry_mac_adresa = Entry(tab2, textvariable=mac_adresa_rucno, width = 25).grid(row=6, column=1)
    label_razmak_a1 = Label(tab2, text=" ").grid(row=7, column=0)

    #label and entry for hardware info
    label_hardware = Label(tab2, text="Hardware: ").grid(row=0, column=3, sticky="e")
    label_model = Label(tab2, text="Model: ").grid(row=1, column=3, sticky="e")
    entry_model = Entry(tab2,textvariable=model_rucno, width = 25).grid(row=1, column=4)
    label_cpu = Label(tab2, text="CPU: ").grid(row=2, column=3, sticky="e")
    entry_cpu = Entry(tab2, textvariable=cpu_rucno, width = 25).grid(row=2, column=4)
    label_ram = Label(tab2, text="RAM: ").grid(row=3, column=3, sticky="e")
    entry_ram = Entry(tab2, textvariable=ram_rucno, width = 25).grid(row=3, column=4)
    label_mrezna_kartica = Label(tab2, text="Mrežna kartica: ").grid(row=4, column=3, sticky="e")
    entry_mrezna_kartica = Entry(tab2, textvariable=mrezna_kartica_rucno, width = 25).grid(row=4, column=4)
    label_graficka_kartica = Label(tab2, text="Grafička kartica: ").grid(row=5, column=3, sticky="e")
    entry_graficka_kartica = Entry(tab2, textvariable=graficka_kartica_rucno, width = 25).grid(row=5, column=4)
    label_glavna_particija = Label(tab2, text="Glavna particija: ").grid(row=6, column=3, sticky="e")
    entry_glavna_particija = Entry(tab2, textvariable=glavna_particija_rucno, width = 25).grid(row=6, column=4)
    label_dodatne_particije = Label(tab2, text="Dodatne particije: ").grid(row=7, column=3, sticky="e")
    entry_dodatne_particije = Entry(tab2, textvariable=dodatne_particije_rucno, width = 25).grid(row=7, column=4)

    #installations
    label_instalacije = Label(tab2, text = "Instalacije: ").grid(row=8, column=0, sticky="e")
    label_os = Label(tab2, text = "OS: ").grid(row=9, column=0, sticky="e")
    entry_os = Entry(tab2, textvariable=os_rucno, width=25).grid(row=9, column=1)

    label_program1 = Label(tab2, text = "Program1: ").grid(row=10, column=0, sticky="e")
    entry_program1_rucno = Entry(tab2, textvariable=program1_rucno, width=25).grid(row=10, column=1)
    label_program2 = Label(tab2, text = "Program2: ").grid(row=11, column=0, sticky="e")
    entry_program2_rucno = Entry(tab2, textvariable=program2_rucno, width=25).grid(row=11, column=1)
    label_program3 = Label(tab2, text = "Program3: ").grid(row=12, column=0, sticky="e")
    entry_program3_rucno = Entry(tab2, textvariable=program3_rucno, width=25).grid(row=12, column=1)
    label_program4 = Label(tab2, text = "Program4: ").grid(row=13, column=0, sticky="e")
    entry_program4_rucno = Entry(tab2, textvariable=program4_rucno, width=25).grid(row=13, column=1)
    label_program5 = Label(tab2, text = "Program5: ").grid(row=14, column=0, sticky="e")
    entry_program5_rucno = Entry(tab2, textvariable=program5_rucno, width=25).grid(row=14, column=1)
    label_program6 = Label(tab2, text = "Program6: ").grid(row=15, column=0, sticky="e")
    entry_program6_rucno = Entry(tab2, textvariable=program6_rucno, width=25).grid(row=15, column=1)
    label_program7 = Label(tab2, text = "Program7: ").grid(row=16, column=0, sticky="e")
    entry_program7_rucno = Entry(tab2, textvariable=program7_rucno, width=25).grid(row=16, column=1)
    label_program8 = Label(tab2, text = "Program8: ").grid(row=17, column=0, sticky="e")
    entry_program8_rucno = Entry(tab2, textvariable=program8_rucno, width=25).grid(row=17, column=1)
    label_program9 = Label(tab2, text = "Program9: ").grid(row=18, column=0, sticky="e")
    entry_program9_rucno = Entry(tab2, textvariable=program9_rucno, width=25).grid(row=18, column=1)
    label_program10 = Label(tab2, text = "Program10: ").grid(row=19, column=0, sticky="e")
    entry_program10_rucno = Entry(tab2, textvariable=program10_rucno, width=25).grid(row=19, column=1)
    label_program11 = Label(tab2, text = "Program11: ").grid(row=20, column=0, sticky="e")
    entry_program11_rucno = Entry(tab2, textvariable=program11_rucno, width=25).grid(row=20, column=1)
    label_program12 = Label(tab2, text = "Program12: ").grid(row=21, column=0, sticky="e")
    entry_program12_rucno = Entry(tab2, textvariable=program12_rucno, width=25).grid(row=21, column=1)
    label_program13 = Label(tab2, text = "Program13: ").grid(row=22, column=0, sticky="e")
    entry_program13_rucno = Entry(tab2, textvariable=program13_rucno, width=25).grid(row=22, column=1)
    label_program14 = Label(tab2, text = "Program14: ").grid(row=23, column=0, sticky="e")
    entry_program14_rucno = Entry(tab2, textvariable=program14_rucno, width=25).grid(row=23, column=1)
    label_program15 = Label(tab2, text = "Program15: ").grid(row=24, column=0, sticky="e")
    entry_program15_rucno = Entry(tab2, textvariable=program15_rucno, width=25).grid(row=24, column=1)


    #notes
    label_napomene = Label(tab2, text = "Napomene: ").grid(row=8, column=3, sticky="e")
    label_napomena1 = Label(tab2, text = "Napomena1: ").grid(row=9, column=3, sticky="e")
    entry_napomena1_rucno = Entry(tab2, textvariable=napomena1_rucno, width=25).grid(row=9, column=4)
    label_napomena2 = Label(tab2, text = "Napomena2: ").grid(row=10, column=3, sticky="e")
    entry_napomena2_rucno = Entry(tab2, textvariable=napomena2_rucno, width=25).grid(row=10, column=4)
    label_napomena3 = Label(tab2, text = "Napomena3: ").grid(row=11, column=3, sticky="e")
    entry_napomena3_rucno = Entry(tab2, textvariable=napomena3_rucno, width=25).grid(row=11, column=4)
    label_napomena4 = Label(tab2, text = "Napomena4: ").grid(row=12, column=3, sticky="e")
    entry_napomena4_rucno = Entry(tab2, textvariable=napomena4_rucno, width=25).grid(row=12, column=4)
    label_napomena5 = Label(tab2, text = "Napomena5: ").grid(row=13, column=3, sticky="e")
    entry_napomena5_rucno = Entry(tab2, textvariable=napomena5_rucno, width=25).grid(row=13, column=4)
    label_napomena6 = Label(tab2, text = "Napomena6: ").grid(row=14, column=3, sticky="e")
    entry_napomena6_rucno = Entry(tab2, textvariable=napomena6_rucno, width=25).grid(row=14, column=4)
    label_napomena7 = Label(tab2, text = "Napomena7: ").grid(row=15, column=3, sticky="e")
    entry_napomena7_rucno = Entry(tab2, textvariable=napomena7_rucno, width=25).grid(row=15, column=4)
    label_napomena8 = Label(tab2, text = "Napomena8: ").grid(row=16, column=3, sticky="e")
    entry_napomena8_rucno = Entry(tab2, textvariable=napomena8_rucno, width=25).grid(row=16, column=4)
    label_napomena9 = Label(tab2, text = "Napomena9: ").grid(row=17, column=3, sticky="e")
    entry_napomena9_rucno = Entry(tab2, textvariable=napomena9_rucno, width=25).grid(row=17, column=4)
    label_napomena10 = Label(tab2, text = "Napomena10: ").grid(row=18, column=3, sticky="e")
    entry_napomena10_rucno = Entry(tab2, textvariable=napomena10_rucno, width=25).grid(row=18, column=4)
    label_napomena11 = Label(tab2, text = "Napomena11: ").grid(row=19, column=3, sticky="e")
    entry_napomena11_rucno = Entry(tab2, textvariable=napomena11_rucno, width=25).grid(row=19, column=4)
    label_napomena12 = Label(tab2, text = "Napomena12: ").grid(row=20, column=3, sticky="e")
    entry_napomena12_rucno = Entry(tab2, textvariable=napomena12_rucno, width=25).grid(row=20, column=4)
    label_napomena13 = Label(tab2, text = "Napomena13: ").grid(row=21, column=3, sticky="e")
    entry_napomena13_rucno = Entry(tab2, textvariable=napomena13_rucno, width=25).grid(row=21, column=4)
    label_napomena14 = Label(tab2, text = "Napomena14: ").grid(row=22, column=3, sticky="e")
    entry_napomena14_rucno = Entry(tab2, textvariable=napomena14_rucno, width=25).grid(row=22, column=4)
    label_napomena15 = Label(tab2, text = "Napomena15: ").grid(row=23, column=3, sticky="e")
    entry_napomena15_rucno = Entry(tab2, textvariable=napomena15_rucno, width=25).grid(row=23, column=4)
    label_napomena16 = Label(tab2, text = "Napomena16: ").grid(row=24, column=3, sticky="e")
    entry_napomena16_rucno = Entry(tab2, textvariable=napomena16_rucno, width=25).grid(row=24, column=4)
    label_napomena17 = Label(tab2, text = "Napomena17: ").grid(row=25, column=3, sticky="e")
    entry_napomena17_rucno = Entry(tab2, textvariable=napomena17_rucno, width=25).grid(row=25, column=4)

    #date
    label_datum = Label(tab2, text = "Datum: ").grid(row=27, column=0, sticky="e")
    entry_datum = Entry(tab2, textvariable=datum_rucno, width=25).grid(row=27, column=1)

    #tab2, employee that did the work
    label_radovi = Label(tab2, text = "Radove izveo: ").grid(row=27, column=3, sticky="e")
    entry_radovi = Entry(tab2,textvariable=radove_izveo_rucno, width=25).grid(row=27, column=4)
    
    #connect to database
    #define proper path for your own machine
    conn = sqlite3.connect('/home/user/database.db')
    #cursor needed for browsing data in the database
    c = conn.cursor()

    #create the table if it does not exist, use it if it exists
    #same field names as for automatic entry
    def stvori_tablicu_rucno():
        c.execute("CREATE TABLE IF NOT EXISTS izvjestaj_tablica(inventurni_broj INTEGER, djelatnik TEXT, "
                  "naziv_racunala TEXT, korisnicko_ime TEXT, lozinka TEXT, mac_adresa TEXT, operacijski_sustav TEXT, "
                  "dodatni_program1 TEXT, dodatni_program2 TEXT, dodatni_program3 TEXT, dodatni_program4 TEXT, "
                  "dodatni_program5 TEXT, dodatni_program6 TEXT, dodatni_program7 TEXT, dodatni_program8 TEXT, "
                  "dodatni_program9 TEXT, dodatni_program10 TEXT, dodatni_program11 TEXT, dodatni_program12 TEXT, "
                  "dodatni_program13 TEXT, dodatni_program14 TEXT, dodatni_program15 TEXT, model TEXT, cpu TEXT, "
                  "ram TEXT, mrezna_kartica TEXT, graficka_kartica TEXT, glavna_particija TEXT, dodatne_particije TEXT, "
                  "napomena1 TEXT, napomena2 TEXT, napomena3 TEXT, napomena4 TEXT, napomena5 TEXT, napomena6 TEXT, "
                  "napomena7 TEXT, napomena8 TEXT, napomena9 TEXT, napomena10 TEXT, napomena11 TEXT, napomena12 TEXT, "
                  "napomena13 TEXT, napomena14 TEXT, napomena15 TEXT, napomena16 TEXT, napomena17 TEXT, datum TEXT, "
                  "radove_izveo TEXT)")

    #data entry, this time get() method gives us the values
    def unesi_podatke_rucno():
        c.execute("INSERT INTO izvjestaj_tablica(inventurni_broj, djelatnik, naziv_racunala, korisnicko_ime, lozinka, "
                  "mac_adresa, operacijski_sustav, dodatni_program1, dodatni_program2, dodatni_program3, dodatni_program4, "
                  "dodatni_program5, dodatni_program6, dodatni_program7, dodatni_program8, dodatni_program9, "
                  "dodatni_program10, dodatni_program11, dodatni_program12, dodatni_program13, dodatni_program14, "
                  "dodatni_program15, model, cpu, ram, mrezna_kartica, graficka_kartica, glavna_particija, "
                  "dodatne_particije, napomena1, napomena2, napomena3, napomena4, napomena5, napomena6, napomena7, "
                  "napomena8, napomena9, napomena10, napomena11, napomena12, napomena13, napomena14, napomena15, "
                  "napomena16, napomena17, datum, radove_izveo) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, "
                  "?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, "
                  "?)", (inventurni_broj_rucno.get(), djelatnik_rucno.get(), naziv_racunala_rucno.get(), korisnicko_ime_rucno.get(), lozinka_rucno.get(),
                        mac_adresa_rucno.get(), operacijski_sustav_rucno.get(), program1_rucno.get(), program2_rucno.get(), program3_rucno.get(), program4_rucno.get(),
                        program5_rucno.get(), program6_rucno.get(), program7_rucno.get(), program8_rucno.get(), program9_rucno.get(),
                        program10_rucno.get(), program11_rucno.get(), program12_rucno.get(), program13_rucno.get(), program14_rucno.get(),
                        program15_rucno.get(), model_rucno.get(), cpu_rucno.get(), ram_rucno.get(), mrezna_kartica_rucno.get(), graficka_kartica_rucno.get(), 
                        glavna_particija_rucno.get(), dodatne_particije_rucno.get(), napomena1_rucno.get(), napomena2_rucno.get(), napomena3_rucno.get(), 
                        napomena4_rucno.get(), napomena5_rucno.get(), napomena6_rucno.get(), napomena7_rucno.get(), napomena8_rucno.get(), napomena9_rucno.get(), 
                        napomena10_rucno.get(), napomena11_rucno.get(), napomena12_rucno.get(), napomena13_rucno.get(), napomena14_rucno.get(), napomena15_rucno.get(),
                        napomena16_rucno.get(), napomena17_rucno.get(), datum_rucno.get(), radove_izveo_rucno.get()))

        #store data and close connection to the database               
        conn.commit()
    
    #execute all defined functions for manual entry
    def rucni_upis(event):
        stvori_tablicu_rucno()
        unesi_podatke_rucno()
        #remove button after click
        pohrani_podatke_rucno.destroy()
        #done message so the user knows it's finished
        gotovo_rucno_label = Label(tab2, text="Pohranjeno!").grid(row=28, column=1)

    #button to store data to database for manual entry
    pohrani_podatke_rucno = Button(tab2, width=15, height=1)
    pohrani_podatke_rucno.config(text='Pohrani u BP!', fg='red', bg='silver')
    pohrani_podatke_rucno.bind("<Button-1>", rucni_upis)
    pohrani_podatke_rucno.grid(row=28, column=1, sticky="e")

#create parent directory  
#Tk() is used for creating the parent directory. Tk class is instanced without any arguments
root = tk.Tk() 
root.geometry("800x650")
#add title for parent directory(root)
root.title("Title") 
#create tab control
#options: Notebook() method takes height, padding and width
#tab creation
#master: tab_control is the parent widget for tabs
#options: Frame() method takes class__, cursor, padding, relief, style, takefocus, height and width.
#explanation: Frame widget acts as a container and is used to group other widgets together 
tab_control = ttk.Notebook(root) 

#adding tabs
tab1 = ttk.Frame(tab_control) 
tab2 = ttk.Frame(tab_control)

#tab1 and tab2 are children widgets of tab_control
#options: add() method takes sticky, state, padding, text, image, compound, underline
#explanation: add() method is in the tk.ttk.Notebook class, used for adding new tabs to the Notebook widget  
tab_control.add(tab1, text ='Automatski unos') 
tab_control.add(tab2, text ='Ručni unos')

#expand: ensures equal division of space between widgets that have a non zero expansion value 
#when the parent widget is expanded
#fill: fill option ensures that the widget occupies the space assigned to it 
#fill=”both” specifies that the widget takes up space on the X and  Y axis 
#fill=”X” specifies that the widget takes up space on the X axis  
#fill=”Y” specifies that the widget takes up space on the Y axis
#pack() method is used to organize widgets into blocks before placement into the parent wwidget
tab_control.pack(expand=1, fill="both") 

#tab1
#enter path to the excel spreadsheet that we need to store to the database, save the entry into a variable in StringVar() format
putanja_automatski_unos = StringVar() 
#explanation: grid() method is used to divide the parent widget into rows and columns
lokacija_entry = Entry(tab1, textvariable=putanja_automatski_unos, width=50).grid(row=0, column=1)

#button for reading the values from the excel spreadsheet after we've entered the path for it
ucitajIzvjestajButton = Button(tab1, width=15, height=1, text='Učitaj izvještaj', fg='red', bg='silver')
ucitajIzvjestajButton.bind("<Button-1>", unos_podataka)
ucitajIzvjestajButton.grid(row=0, column=0)

#tab2
#button to call all defined functions for manual entry
rucni_unos_button = Button(tab2, width=15, height=1, text='Ručni unos', fg='red', bg='silver')
rucni_unos_button.bind("<Button-1>", rucni_unos)
rucni_unos_button.grid(row=0, column=0)

#run program
#mainloop() acts as an infinite loop and is used for running the program
root.mainloop()   